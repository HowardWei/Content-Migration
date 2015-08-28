__author__ = 'Howard Wei / Beyin Abraha'

from openpyxl import load_workbook
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import Select
from bs4 import BeautifulSoup
from urllib.parse import urlparse
import pyperclip
import time
import re
import urllib.request

# This information needs to be changed for each school
sharpURL = "http://union.ss8.sharpschool.com/"
oldURL = "http://www.westirondequoit.org/"

# Excel sheet information
filePath = "C:\\Users\\hannan.wei\\Downloads\\"
username = "han.nan.wei"    
fileName = "WestIrondequoitCentralSchoolDistrict.xlsx"
sheetName = "District"
#----------------------------------------------------


homeID = ""
workbook = Workbook()
stackURL = ["","","",""]

#def fetch():
#    username = input("Enter the login username: ")
#    fileName = input("Enter the filename of the excel document: ")
#    sheetName = input("Enter the filename of the excel document: ")
#    dimY = input("Enter the number of rows: ")
#    ++dimY


def login(driver):
    driver.find_element_by_id("ctl00_lnkGateway").click()
    WebDriverWait(driver, 10).until(ec.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_txtUsername")))
    input = driver.find_element_by_id('ctl00_ContentPlaceHolder1_txtUsername')
    input.send_keys(username)
    input = driver.find_element_by_id("ctl00_ContentPlaceHolder1_txtPassword")
    input.send_keys('welcome')
    driver.find_element_by_id("ctl00_ContentPlaceHolder1_btnLogin").click()
    print("stackURL.append(driver.current_url) ->" + driver.current_url)
    stackURL[0] = driver.current_url


#---- TraversePages function used to walk through the excel and fetch ----#
#---- workbook document and printout the mapping of the excel document -----#


def FileSetup(fileName, driver):
    workbook = load_workbook(filename = filePath + fileName)
    global excelSheet
    excelSheet = workbook.get_sheet_by_name(sheetName)
    

#---- GetHomeID function not used in this implementation ----#
    
#def GetHomeID(driver):
#    WebDriverWait(driver, 10).until(ec.presence_of_element_located((By.CSS_SELECTOR, "span.text.expandTop")))
#    driver.find_element_by_css_selector("span.text.expandTop").click()
#    driver.find_element_by_css_selector("#ctl00_ContentPlaceHolder1_ctl00_ctl00_menu_m0_m0 > span.text").click()
#    homeURL = driver.current_url
#    print (homeURL)


def GetContentCell(cell):
    
    domain = urlparse(oldURL)
    print(domain.netloc)
    
    listCell = list(cell)
    listCell[0] = "D"
    cell = "".join(listCell)
    print(cell)
    
    return cell

def DeterminePage(driver, cell):
    
    global pageURL
    pageURL = excelSheet[GetContentCell(cell)].value
    
    if(pageURL == None):
        return "Content"
    
    print(urlparse(pageURL).netloc)
    
    testDomain = urlparse(pageURL)
    
    if(testDomain.netloc == urlparse(oldURL).netloc):
        if((".doc" or ".docx" or ".jpg" or ".pdf" or ".mov" or ".wma"
        ".swf" or ".mp3" or ".mp4" or ".php" or ".ppt" or ".pptx") in pageURL):
            return "File"
        else:
            return "Content"
    else:
        return "External"


def CreatePages(driver, excelSheet, stackURL):
    
    global level
    level = 1
    notDone = True
    y = 2
    global x 
    x = 0
    
    while notDone:
        print("\n NEW PAGE \n")
        print(excelSheet[array[x] + str(y)].value)
        print(level)
        print(stackURL)
        
        CreatePage(driver, stackURL, array[x] + str(y))
        
        if (x < 2 and excelSheet[array[level] + str(y + 1)].value != None):
            level += 1
            x += 1
            y += 1
            driver.get(stackURL[x])
        elif (excelSheet[array[x] + str(y + 1)].value != None):
            y += 1
            driver.get(stackURL[x])
        elif (excelSheet[array[x - 1] + str(y + 1)].value != None):
            level -= 1
            x -= 1
            y += 1
            driver.get(stackURL[x])
        elif (excelSheet[array[x - 2] + str(y + 1)].value != None):
            level -= 2
            x -= 2
            y += 1
            driver.get(stackURL[x])
        else:
            print ("End of excel sheet")
            notDone = False


def CreatePage(driver, stackURL, cell):
    
    pageName = excelSheet[cell].value
    typePage = DeterminePage(driver, cell)  
    
    testID = "ctl00_ContentPlaceHolder1_ctl00_ctl00_menu_m2"
    #testID = GetID(driver, "ctl00_ContentPlaceHolder1_ctl00_ctl00_menu_m2", True)
    
    WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.ID, testID)))
    driver.find_element_by_id(testID).click()
    
    if(typePage == "Content"): num = "6"
    else: num = "7"
    
    testID = "ctl00_ContentPlaceHolder1_ctl00_ctl00_menu_m2_m" + num
    #testID = GetID(driver,"ctl00_ContentPlaceHolder1_ctl00_ctl00_menu_m2_m" + num, True)
    
    WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.ID, testID)))
    driver.find_element_by_id(testID).click()
    
    WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.ID, GetID(driver, "ctl00_ContentPlaceHolder1_ctl00_txtTitle", False))))    
    input = driver.find_element_by_id(GetID(driver, "ctl00_ContentPlaceHolder1_ctl00_txtTitle", False))
    input.send_keys(pageName)
    
    print(typePage)
    if(typePage == "Content"):
        CreateContentPage(driver, stackURL, cell)
    elif(typePage == "External"):
        CreateLinkPage(driver, stackURL, cell)
    elif(typePage == "File"):
        CreateFilePage(driver, stackURL, cell)


def CreateContentPage(driver, stackURL, cell):
    
    created = True
    pageIT = 0
    pageName = excelSheet[cell].value
    testID = GetID(driver, "ctl00_ContentPlaceHolder1_ctl00_txtTitle", False)
    
    input = driver.find_element_by_id(testID)
    input.send_keys(Keys.RETURN)
    
    while(created):
        try:
            WebDriverWait(driver, 5).until(ec.visibility_of_element_located((By.CLASS_NAME, "reMode_html")))
            created = False
        except:
            pageIT += 1
            input = driver.find_element_by_id(GetID(driver, "ctl00_ContentPlaceHolder1_ctl00_txtTitle", False))
            input.clear()
            input.send_keys(pageName + " (" + str(pageIT) + ")")
            input.send_keys(Keys.RETURN)
        
    if(excelSheet[GetContentCell(cell)].value != None):
        content = GrabContent(driver, cell)
        pyperclip.copy(content)

        if(content != None):
            driver.find_element_by_class_name("reMode_html").click()
            
            textbox = driver.find_elements_by_tag_name("iframe")[1]
            textbox.send_keys(Keys.CONTROL + "v")
            
            WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.ID, "loadBtn")))
            driver.find_element_by_id("loadBtn").click()
            print("clicked loadBtn")
            WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.ID, "stripBtn")))
            driver.find_element_by_id("stripBtn").click()
            print("clicked stripBtn")        
            WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.ID, "startBtn")))
            driver.find_element_by_id("startBtn").click()
            print("clicked startBtn")
            
            try:
                WebDriverWait(driver, 999).until(ec.element_to_be_clickable((By.ID, "startBtn")))
            except:
                print ("An error occured, skipping page. \n Cell ->" + cell)
    
    
    testID = GetID(driver, "ctl00_ContentPlaceHolder1_ctl00_ibPublishBottom", False)
    WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.ID, testID)))
    driver.find_element_by_id(testID).click()
    
    testID = GetID(driver, "ctl00_ContentPlaceHolder1_ctl00_btnYes", False)
    WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.ID, testID)))
    driver.find_element_by_id(testID).click()
    
    stackURL[level] = driver.current_url


def CreateLinkPage(driver, stackURL, cell):
    input = driver.find_element_by_id(GetID(driver, "ctl00_ContentPlaceHolder1_ctl00_txtUrl", False))
    input.send_keys(pageURL)
    
    input = driver.find_element_by_id(GetID(driver, "ctl00_ContentPlaceHolder1_ctl00_hplGetName", False))
    input.click()
    
    select = Select(driver.find_element_by_id(GetID(driver, "ctl00_ContentPlaceHolder1_ctl00_ddlProtocol", False)))
    select.select_by_value(urlparse(pageURL).scheme + ":\/\/")
    
    input = driver.find_element_by_id(GetID(driver, "ctl00_ContentPlaceHolder1_ctl00_btnSubmit", False))
    input.click()
    
    stackURL[level] = driver.current_url


def ParseContent(cell, content):
    listOfImages = content.findAll('img')
    
    for img in listOfImages:
        imgURL = img["src"]
        tempURL = excelSheet[cell].value
        print (oldURL + imgURL)
        img["src"] = oldURL + imgURL
    
    content = str(content)
    content.replace("","")
    
    return content


def GrabContent(driver, cell):
    url = excelSheet[GetContentCell(cell)].value
    
    req = urllib.request.Request(url)
    html = urllib.request.urlopen(req).read()
    soup = BeautifulSoup(html, "html.parser")
    content = soup.find("div", {"id": "edlThemeMainsectionZoneFrame"})
    content = ParseContent(cell, content)
    return content


def CreateFilePage(driver, stackURL, cell):
    input = driver.find_element_by_id(GetID(driver, "ctl00_ContentPlaceHolder1_ctl00_rblTypes_1", False))
    input.click()
    
    input = driver.find_element_by_id(GetID(driver, "ctl00_ContentPlaceHolder1_ctl06_txtUrl", False))
    file = GetContentCell(cell)    
    input.send_keys(excelSheet(file))
    
    input.find_element_by_id("chkBtn")
    input.click()
    WebDriverWait(driver, 999).until(ec.element_to_be_clickable((By.ID, "chkBtn")))
    
    input.find_element_by_id(GetID(driver, "ctl00_ContentPlaceHolder1_ctl06_btnSubmit", False))
    input.click()
    
    return

def GetID(driver, baseID, header):
        
    for ID in range (0, 10): 
        
        try:
            
            if header:
                testID = baseID.replace("ContentPlaceHolder1_ctl00_ctl00", "ContentPlaceHolder1_ctl0" + str(ID) + "_ctl0" + str(ID))
            else:
                testID = baseID.replace("ContentPlaceHolder1_ctl00", "ContentPlaceHolder1_ctl0" + str(ID))
            
            if (ID == 0):
                WebDriverWait(driver, 0.5).until(ec.visibility_of_element_located((By.ID, testID)))
            
            element = driver.find_element_by_id(testID)
            if element.is_displayed():
                break
            
        except:
            #print ("This ID cannot be found on the page: " + testID)
            continue
    
    #print(testID)
    testID = "".join(testID)
    
    return testID


class ExcelSkeleton:

    #fetch()
    global array 
    array = ["A", "B", "C"]
    fp = webdriver.FirefoxProfile("Howard")
    driver = webdriver.Firefox(firefox_profile=fp)
    driver.get(sharpURL)
    login(driver)
    
    FileSetup(fileName, driver)
    CreatePages(driver, excelSheet, stackURL)
    
    driver.close()